import io
import os
from os import path as relative_path
import shutil
import logging
import datetime
import calendar
import functools
import openpyxl
import pylint.lint

path = 'mini_project\mini_project_part1.py'
pylint_opts = ['--disable=line-too-long','--disable=trailing-whitespace','--disable=logging-format-interpolation','--disable=consider-using-f-string', path]

logging.basicConfig(filename="mini_project\logs.log", filemode='w',level = logging.DEBUG, format = '%(asctime)s:[%(levelname)-8s]: %(message)s', datefmt='%Y-%m-%d %H:%M:%S')


def is_processed(file_path, filelst):
    if isinstance(filelst,str):
        filelst = open(filelst, "r+")
    else:
        filelst = open(filelst.name, "r+")
    lst = filelst.read()
    found = False
    if file_path in lst:
        found = True
    filelst.close()
    return found


def marked_as_processed(file_path, filelst):
    if isinstance(filelst,str):
        filelst = open(filelst, "a+")
    else:
        filelst = open(filelst.name, "a+")
    filelst.write(file_path + "\n")
    filelst.close()


def verify_file_name(path, extension):
    flag = True
    if relative_path.isfile(path) is False:
        #logging.error("Provided path {0} is not a file program ended".format(path))
        logging.error(f"Provided path {path} is not a file program ended")
        flag = False
    if extension != ".xlsx":
        logging.critical("File type {0} is not supported".format(extension))
        flag = False
    if relative_path.exists(path) is False or flag is False:
        logging.critical("File {0} not found program ended".format(path))
    else:
        logging.info("File {0} successfully found".format(path))
        return flag
    return flag


def parse_file_name(file):
    replacements = ('_')
    file = functools.reduce(lambda s, sep: s.replace(sep, ' '), replacements, file)
    return file.split()
    

def is_valid_excel_file(wb_obj):
    excepted_sheet_names = ['Summary Rolling MoM','VOC Rolling MoM','Monthly Verbatim Statements']
    tab_amt = len(wb_obj.sheetnames)
    if tab_amt != len(excepted_sheet_names):
        return False
    for tab in wb_obj.sheetnames:
        if tab in excepted_sheet_names == False:
            return False
    return True


def get_month_year(parsed):#list of strings as input
    month_year = [None,None]
    for text in parsed:
        if text.isnumeric():
            month_year[1] = int(text) # gets year
            logging.info(f"Found year from file name {month_year[1]}")
            if isinstance(month_year[0], int):
                #can return because year is given after month in the formating described
                return month_year[0], month_year[1]
        try:
            month = datetime.datetime.strptime(text, "%B").month
            month_year[0] = month
            logging.info("Found month from file name '{0}' converting to month number {1} for later use".format(text, month_year[0]))

        except:
            logging.warning("Still searching for year and month from file name")

    logging.critical("could not find month and year program ended")
    return False


#function does a smart scan
def get_month_year_cell_positions(sheet_obj, file_month, file_year):
    #list of cell coordionates that met description
    matching_datetime_cells = []
    logging.debug("Starting search for all locations of rows with information about report on month {0} and year {1}".format(file_month, file_year))
    for row in sheet_obj:
        for column in row:
            if isinstance(column.value, datetime.datetime):
                if column.value.month == file_month and column.value.year == file_year:#gets the exact datetime here
                    cell_coordinates = (column.row, column.column)
                    logging.info("Found cell with approriate dating at cell_coordinates {0} with value {1}".format(cell_coordinates, column.value))
                    matching_datetime_cells.append(cell_coordinates)
    amt_of_matches = len(matching_datetime_cells)
    logging.info("Found {0} cells with approriate dating".format(amt_of_matches))
    return matching_datetime_cells


def get_row_information(sheet_obj, row, column):#assumes datetime can be in any column bu the months data is towards is right as specified 
    row_length = len(sheet_obj[row])
    row_info = []
    seen_datetime = False # used so if two month cells are in same row so their values are seperated
    for x in range(0, row_length - column + 1):
        cell_value = sheet_obj.cell(row, column + x).value
        cell_column_name = sheet_obj.cell(1, column + x).value#can use 1 since the column name with be on top
        if isinstance(cell_value,datetime.datetime):
            if seen_datetime is False:
                row_info.append([cell_value, cell_column_name])
                seen_datetime = True
            else:
                return
        elif(cell_value is None or cell_column_name is None) is False:
            if isinstance(cell_value,float) is False:
                row_info.append([cell_value, cell_column_name])
            else:
                row_info.append([cell_value * 100, cell_column_name])
    return row_info


def print_row_in_priority_order(row):
    displaypriority = {
    "Calls Offered" : 0,
    "Abandon after 30s" : 1,
    "FCR" : 2,
    "DSAT" : 3,
    "CSAT" : 4
    }
    logging.info(row[0][0].strftime("Values being displayed for %B"))
    row.remove(row[0])

    sorted(row, key=lambda x: displaypriority[x[1].strip()], reverse=False)#sort information based on a dictionary lookup key and  example given
    for cell in row:
        cell[1] = cell[1].strip() #sanitize display for output
        if isinstance(cell[0],float):
            logging.info("{} : {}%".format(cell[1], cell[0])) #for display percetages
        else:
            logging.info("{} : {}".format(cell[1], cell[0])) #for non percentages


def iterate_column(sheet_obj, column_num):#generator
    x = 1
    while x < sheet_obj.max_row:
        yield sheet_obj.cell(row = x, column = column_num)
        x+=1


def score_review(promoter_type, score):
    if promoter_type == "Promoters":
        return (lambda x: "good because greater than 200" if x>200  else "bad because lesser than 200")(score)
    elif promoter_type == "Passives" or promoter_type == "Dectractors":
        return (lambda x: "good because greater than 100" if x>100  else "bad because lesser than 100")(score)
    else:
        return "not valid promoter type given"

def find_performance_scores(sheet_obj,column):
    desired__performance_score_metrics = ["Promoters", "Passives", "Dectractors"]
    gen = iterate_column(sheet_obj, column)
    x = 1 
    try:
        while x < sheet_obj.max_row:
            possible_desired_value = next(gen)
            row_pos = possible_desired_value.row
            row_title  = sheet_obj[row_pos][0].value
            if isinstance(row_title,str):
                row_title = row_title.split(" ")
                for word in row_title:
                    if word in desired__performance_score_metrics:
                        logging.info("{} : {} : {}".format(word, possible_desired_value.value, score_review(word,possible_desired_value.value)))#is desired of reaches here
            x+=1
    except:
        logging.info("processing done for info on calender info given {}".format(sheet_obj.cell(row =1, column = column).value)) 
        gen.close()

def get_performance_scores(sheet_obj, file_name_month, file_name_year):#month and year are ints
    for column_header in sheet_obj[1]:
        cell = column_header.value
        if isinstance(cell,datetime.datetime):
            if cell.month == file_name_month and cell.year == file_name_year:#gets the exact datetime here
                find_performance_scores(sheet_obj, column_header.column)
                return True
        elif isinstance(cell,str):
            if cell == calendar.month_name[file_name_month]:
                find_performance_scores(sheet_obj, column_header.column)
                return True

    logging.critical("Could not find suitable month and year combination on tab {0} invalid file".format(sheet_obj.title))         
    return False


#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

logging.debug("Start of program mini_project")
archive_directory = os.fsencode("C:\\Users\\mskwa_000\\Documents\\GitHub\\smoothstack_assignments\\mini_project\\archive_directory").decode('utf-8')
search_directory =  os.fsencode("C:\\Users\\mskwa_000\\Documents\\GitHub\\smoothstack_assignments\\mini_project\\search_directory").decode('utf-8')
error_directory =   os.fsencode("C:\\Users\\mskwa_000\\Documents\\GitHub\\smoothstack_assignments\\mini_project\\error_directory").decode('utf-8')
filelst = "C:\\Users\\mskwa_000\\Documents\\GitHub\\smoothstack_assignments\\mini_project\\filelst.lst"


try:#creates file if not already created
    filelst = open(filelst,'x')
    filelst.close()
except:
    pass
#datetime.strftime('%Y-%m-%d )
files = next(os.walk(search_directory), (None, None, []))[2]  # [] if no file
# try:
#     files.append(input("Enter complete file path here: "))
# except:
#     pass
for file_path in files:
    file_path = search_directory + "\\" + file_path
    if is_processed(file_path, filelst):
        logging.error("File {} already processed moving to error directory {}".format(file_path, error_directory))
        shutil.move(file_path, error_directory)
        continue
    logging.info("Starting processing of file {}".format(file_path))

    logging.info("Acessing file {}".format(file_path))
    file_name, extension = os.path.splitext(file_path)
    with open(file_path, 'rb') as f:
        file = io.BytesIO(f.read())

    if verify_file_name(file_path, extension) == False:
        logging.critical("file name {0} could not be verify program ended".format(file_path))
        logging.info("file name {0} moved to error directory {1}".format(file_path, error_directory))
        shutil.move(file_path, error_directory)
        continue
    logging.info("file name {0} could be verified program continuing".format(file_path))

    logging.debug("Starting to parse file name {0}".format(file_path))
    parse = parse_file_name(file_name)
    logging.info("Successfully parsed file_name {0} and is of supported type {1}".format(file_path, extension))

    logging.debug("Starting to load excel worksheet from file {}".format(file_path))
    wb_obj = openpyxl.load_workbook(file, read_only="True")

    #checks to see if excel file is missing a tab    
    if is_valid_excel_file(wb_obj) is False:
        logging.info("file {} is not a valid excel file moving to error_directory {}".format(file_path, error_directory))
        marked_as_processed(file_path, filelst)
        shutil.move(file_path, error_directory)
        continue
    logging.info("file {} is validated excel file".format(file_path))

    logging.info("Successfully loaded excel workbook")
    
    logging.debug("Starting search for month and year from parse of file name {0}".format(file_path))
    
    try:
        file_name_month, file_name_year = get_month_year(parse)
    except:
        logging.info("file did not have valid information moving to error directory".format(error_directory))
        marked_as_processed(file_path, filelst)
        shutil.move(file_path, error_directory)
        continue
    logging.info("Completed search for month and year from file name {0} : file_name_month = {1}, file_name_year = {2}".format(file_path, file_name_month, file_name_year))

    logging.info("Acessing excel workbook tab 'Summary Rolling MoM'")
    sheet_obj = wb_obj['Summary Rolling MoM']
    cell_positions = get_month_year_cell_positions(sheet_obj, file_name_month, file_name_year)#return list of (row,column)
    if len(cell_positions) == 0:
        logging.info("Since could not find any appropriate imformation stopping processing of file".format(file_path))
        marked_as_processed(file_path, filelst)
        shutil.move(file_path, error_directory)
        continue
    rows_info = []
    
    for match in cell_positions:# to be able to handle all instances that can fit the described time frame
        rows_info.append(get_row_information(sheet_obj, match[0], match[1]))
    for row in rows_info:
        print_row_in_priority_order(row)

    logging.info("Acessing excel workbook tab 'VOC Rolling MoM'")
    sheet_obj = wb_obj['VOC Rolling MoM']
    logging.info("Loading tab 'VOC Rolling MoM'")
    if get_performance_scores(sheet_obj, file_name_month, file_name_year):
        logging.info("Found proper information from tab {}".format(sheet_obj.title))
    else:
        logging.info("Did not find proper information from tab {}".format(sheet_obj.title))
    logging.info("Processing for file {} finished moving to archive directory {}".format(file_path, archive_directory))  
    marked_as_processed(file_path,filelst)
    shutil.move(file_path,archive_directory)
logging.info("Program finished processing all files in search directory {}".format(search_directory))
pylint.lint.Run(pylint_opts)
